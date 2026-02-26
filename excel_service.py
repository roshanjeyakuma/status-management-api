# 1. Standardize Project Name
def format_project_name(name: str):
    return name.strip().lower().replace(" ", "_")

# 2. Create Excel File If Not Exists
import os
import pandas as pd
from openpyxl import load_workbook

FILE_NAME = "company_status.xlsx"

def ensure_file_exists():
    if not os.path.exists(FILE_NAME):
        df = pd.DataFrame()
        df.to_excel(FILE_NAME, index=False)

# 3. Create Project Sheet
    
def create_project(project_name: str):
    project = format_project_name(project_name)
    ensure_file_exists()

    try:
        book = load_workbook(FILE_NAME)
        if project in book.sheetnames:
            return "Project already exists"
    except:
        pass

    df = pd.DataFrame(columns=["Date", "Status Description"])

    with pd.ExcelWriter(FILE_NAME, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name=project, index=False)

    return "Project created successfully"

# 4. Add Update to Project Sheet

def add_update(project_name: str, date: str, status: str):
    project = format_project_name(project_name)
    ensure_file_exists()

    book = load_workbook(FILE_NAME)

    if project not in book.sheetnames:
        return "Project does not exist"

    df = pd.read_excel(FILE_NAME, sheet_name=project)

    new_row = {
        "Date": date,
        "Status Description": status
    }

    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

    with pd.ExcelWriter(FILE_NAME, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=project, index=False)

    return "Update added successfully"

# 5. Get Updates for a Specific Date

def get_updates(project_name: str, date: str):
    project = format_project_name(project_name)

    book = load_workbook(FILE_NAME)

    if project not in book.sheetnames:
        return "Project does not exist"

    df = pd.read_excel(FILE_NAME, sheet_name=project)

    filtered = df[df["Date"] == date]

    if filtered.empty:
        return []

    return filtered["Status Description"].tolist()

# 6. Edit Existing Update

def edit_update(project_name: str, date: str, entry_number: int, new_text: str):
    project = format_project_name(project_name)

    df = pd.read_excel(FILE_NAME, sheet_name=project)

    filtered = df[df["Date"] == date]

    if filtered.empty:
        return "No entries for that date"

    actual_index = filtered.index[entry_number - 1]

    df.loc[actual_index, "Status Description"] = new_text

    with pd.ExcelWriter(FILE_NAME, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=project, index=False)

    return "Entry updated successfully"